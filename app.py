import cv2
import sqlite3
import threading
import time
import hashlib
import os
import numpy as np
from datetime import datetime
from collections import deque
from dotenv import load_dotenv
from flask import (Flask, Response, render_template, request, redirect,
                   url_for, session, jsonify, stream_with_context, send_file)
import torch
from ultralytics import YOLO
import google.generativeai as genai

# Fix for PyTorch 2.6+ weights_only security
torch.serialization.add_safe_globals(["ultralytics.nn.tasks.DetectionModel"])

load_dotenv()
genai.configure(api_key=os.getenv("gemini_key"))
gemini = genai.GenerativeModel("gemini-2.5-flash")

app = Flask(__name__)
app.secret_key = "monitor_secret_2024"
model = YOLO("yolov8n.pt", task="detect")
DB = "monitor.db"
RECORDINGS_DIR = "recordings"
os.makedirs(RECORDINGS_DIR, exist_ok=True)

cameras = {}

COCO_CLASSES = {0: "person", 1: "bicycle", 2: "car", 3: "motorcycle", 5: "bus", 7: "truck"}


# ── Helpers ───────────────────────────────────────────────────────────────────
def hash_pw(pw):
    return hashlib.sha256(pw.encode()).hexdigest()


def get_db():
    con = sqlite3.connect(DB)
    con.row_factory = sqlite3.Row
    return con


def init_db():
    with get_db() as con:
        con.executescript("""
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                password TEXT NOT NULL,
                role TEXT NOT NULL DEFAULT 'user'
            );
            CREATE TABLE IF NOT EXISTS cameras (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                url  TEXT NOT NULL,
                active INTEGER DEFAULT 1
            );
            CREATE TABLE IF NOT EXISTS detections (
                id        INTEGER PRIMARY KEY AUTOINCREMENT,
                cam_id    INTEGER,
                cam_name  TEXT,
                details   TEXT,
                timestamp TEXT
            );
            CREATE TABLE IF NOT EXISTS recordings (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                cam_id     INTEGER,
                cam_name   TEXT,
                filename   TEXT,
                start_time TEXT,
                end_time   TEXT
            );
        """)
        if not con.execute("SELECT 1 FROM users WHERE username='admin'").fetchone():
            con.execute("INSERT INTO users (username,password,role) VALUES (?,?,?)",
                        ("admin", hash_pw("admin123"), "admin"))


def get_dominant_color(crop):
    if crop is None or crop.size == 0:
        return "unknown"
    try:
        pixels = np.float32(crop.reshape(-1, 3))
        _, _, centers = cv2.kmeans(pixels, 1, None,
                                   (cv2.TERM_CRITERIA_EPS + cv2.TERM_CRITERIA_MAX_ITER, 10, 1.0),
                                   3, cv2.KMEANS_PP_CENTERS)
        b, g, r = [int(c) for c in centers[0]]
        if r > 180 and g < 100 and b < 100: return "red"
        if r < 80 and g > 160 and b < 100:  return "green"
        if r < 80 and g < 100 and b > 180:  return "blue"
        if r > 180 and g > 180 and b < 80:  return "yellow"
        if r < 50  and g < 50  and b < 50:  return "black"
        if r > 200 and g > 200 and b > 200: return "white"
        if r > 160 and g > 100 and b < 60:  return "orange"
        if r > 100 and g < 80  and b > 140: return "purple"
        if r > 150 and g > 120 and b > 100: return "grey"
        return "mixed"
    except:
        return "unknown"


def estimate_gender(crop):
    if crop is None or crop.size == 0:
        return "unknown"
    h, w = crop.shape[:2]
    if h < 60:
        return "unknown"
    # Upper body color vs lower body color heuristic
    upper = crop[:h//2, :]
    lower = crop[h//2:, :]
    upper_color = get_dominant_color(upper)
    lower_color = get_dominant_color(lower)
    aspect = h / max(w, 1)
    if lower_color in ("black", "grey", "blue") and aspect > 2.0:
        return "likely male"
    if lower_color in ("white", "red", "mixed") and aspect > 2.2:
        return "likely female"
    return "unknown"


# ── Camera thread ─────────────────────────────────────────────────────────────
def camera_worker(cam_id, url, name):
    state = cameras[cam_id]
    cap = cv2.VideoCapture(url)
    cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)

    fps = 25
    state["fps"] = fps
    state["buffer"] = deque(maxlen=fps * 60)

    prev_details = ""
    last_annotated = None
    last_detections = []

    # ── Detection runs in its own thread so video never blocks ──
    def detect_loop():
        nonlocal last_annotated, last_detections
        while state["active"]:
            raw = state.get("_raw_frame")
            if raw is None:
                time.sleep(0.02)
                continue
            try:
                results = model(raw, conf=0.4, iou=0.45, verbose=False)[0]
            except:
                time.sleep(0.02)
                continue

            detections = []
            annotated = raw.copy()
            for b in results.boxes:
                cls_id = int(b.cls[0])
                if cls_id not in COCO_CLASSES:
                    continue
                cls_name = COCO_CLASSES[cls_id]
                x1, y1, x2, y2 = map(int, b.xyxy[0])
                x1 = max(0, min(x1, raw.shape[1]))
                y1 = max(0, min(y1, raw.shape[0]))
                x2 = max(0, min(x2, raw.shape[1]))
                y2 = max(0, min(y2, raw.shape[0]))
                if x2 <= x1 or y2 <= y1:
                    continue
                crop = raw[y1:y2, x1:x2]
                color = get_dominant_color(crop)
                if cls_name == "person":
                    gender = estimate_gender(crop)
                    detail = f"Person ({gender}, {color})"
                    box_color = (34, 197, 94)
                elif cls_name in ("car", "bus", "truck"):
                    detail = f"{cls_name.capitalize()} ({color})"
                    box_color = (37, 99, 235)
                else:
                    detail = f"{cls_name.capitalize()} ({color})"
                    box_color = (239, 68, 68)
                detections.append(detail)
                cv2.rectangle(annotated, (x1, y1), (x2, y2), box_color, 2)
                cv2.putText(annotated, detail, (x1 + 3, max(y1 - 5, 10)),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.42, box_color, 1)

            last_detections = detections
            last_annotated = annotated

    threading.Thread(target=detect_loop, daemon=True).start()

    while state["active"]:
        ret, frame = cap.read()
        if not ret:
            time.sleep(0.3)
            cap = cv2.VideoCapture(url)
            cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)
            continue

        ts = datetime.now()
        h, w = frame.shape[:2]
        # 640 width — good balance of quality and speed
        frame_resized = cv2.resize(frame, (640, int(h * 640 / w)))

        # Feed latest frame to detector
        state["_raw_frame"] = frame_resized
        state["buffer"].append((ts, frame_resized.copy()))

        # Use latest annotated frame if available, else raw
        display = last_annotated if last_annotated is not None else frame_resized
        display = display.copy()

        # Overlay summary
        count = len(last_detections)
        summary = f"Objects: {count}"
        cv2.putText(display, summary, (10, 25), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 0), 2)
        cv2.putText(display, summary, (10, 25), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 1)

        # REC indicator
        if state.get("recording"):
            fw = display.shape[1]
            cv2.circle(display, (fw - 20, 20), 6, (0, 0, 255), -1)
            cv2.putText(display, "REC", (fw - 52, 23),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.45, (0, 0, 255), 1)

        ts_str = ts.strftime("%Y-%m-%d %H:%M:%S")
        state["frame"] = display
        state["detections"] = last_detections
        state["time"] = ts_str

        # DB log on change
        details_str = ", ".join(last_detections)
        if last_detections and details_str != prev_details:
            try:
                with get_db() as con:
                    con.execute(
                        "INSERT INTO detections (cam_id,cam_name,details,timestamp) VALUES (?,?,?,?)",
                        (cam_id, name, details_str, ts_str)
                    )
                prev_details = details_str
            except:
                pass

        # Record
        if state.get("recording") and state.get("writer"):
            try:
                state["writer"].write(display)
            except:
                pass

    cap.release()
    if state.get("writer"):
        state["writer"].release()


def start_camera(cam_id, url, name):
    if cam_id in cameras and cameras[cam_id]["active"]:
        return
    cameras[cam_id] = {
        "url": url, "name": name,
        "frame": None, "detections": [], "time": "",
        "lock": threading.Lock(), "active": True,
        "recording": False, "writer": None,
        "buffer": deque(maxlen=1200),
        "rec_start": None, "rec_filename": None
    }
    threading.Thread(target=camera_worker, args=(cam_id, url, name), daemon=True).start()


def stop_camera(cam_id):
    if cam_id in cameras:
        cameras[cam_id]["active"] = False
        cameras.pop(cam_id, None)


def load_cameras_from_db():
    with get_db() as con:
        rows = con.execute("SELECT id,name,url FROM cameras WHERE active=1").fetchall()
    for r in rows:
        start_camera(r["id"], r["url"], r["name"])


# ── Auth ──────────────────────────────────────────────────────────────────────
def logged_in():
    return "user" in session


def is_admin():
    return session.get("role") == "admin"


def login_required(f):
    from functools import wraps
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not logged_in():
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return wrapper


def admin_required(f):
    from functools import wraps
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not logged_in() or not is_admin():
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return wrapper


# ── Routes ────────────────────────────────────────────────────────────────────
@app.route("/", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        username = request.form["username"]
        password = hash_pw(request.form["password"])
        expected_role = request.form.get("expected_role", "user")
        with get_db() as con:
            user = con.execute(
                "SELECT * FROM users WHERE username=? AND password=?",
                (username, password)
            ).fetchone()
        if user:
            if user["role"] != expected_role:
                error = f"This account is not registered as {expected_role}."
            else:
                session["user"] = user["username"]
                session["role"] = user["role"]
                return redirect(url_for("admin" if user["role"] == "admin" else "dashboard"))
        else:
            error = "Invalid username or password"
    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/dashboard")
@login_required
def dashboard():
    with get_db() as con:
        cams = con.execute("SELECT * FROM cameras WHERE active=1").fetchall()
        total = con.execute("SELECT COUNT(*) as c FROM detections").fetchone()["c"]
        recent = con.execute("SELECT * FROM detections ORDER BY id DESC LIMIT 30").fetchall()
    return render_template("dashboard.html",
                           cams=cams, total=total, recent=recent,
                           user=session["user"], role=session["role"])


@app.route("/admin", methods=["GET", "POST"])
@admin_required
def admin():
    msg = None
    if request.method == "POST":
        action = request.form.get("action")
        if action == "add_cam":
            name = request.form["name"].strip()
            ip_or_url = request.form["ip"].strip()
            # Support both IP and full URLs
            if ip_or_url.startswith("http"):
                url = ip_or_url
            else:
                url = f"http://{ip_or_url}:8080/video"
            with get_db() as con:
                cur = con.execute("INSERT INTO cameras (name,url) VALUES (?,?)", (name, url))
                cam_id = cur.lastrowid
            start_camera(cam_id, url, name)
            msg = f"Camera '{name}' added."
        elif action == "delete_cam":
            cam_id = int(request.form["cam_id"])
            stop_camera(cam_id)
            with get_db() as con:
                con.execute("UPDATE cameras SET active=0 WHERE id=?", (cam_id,))
            msg = "Camera removed."
        elif action == "add_user":
            uname = request.form["uname"].strip()
            pw = request.form["pw"].strip()
            role = request.form["role"]
            try:
                with get_db() as con:
                    con.execute("INSERT INTO users (username,password,role) VALUES (?,?,?)",
                                (uname, hash_pw(pw), role))
                msg = f"User '{uname}' created."
            except sqlite3.IntegrityError:
                msg = "Username already exists."
        elif action == "delete_user":
            uid = int(request.form["uid"])
            with get_db() as con:
                con.execute("DELETE FROM users WHERE id=? AND username!='admin'", (uid,))
            msg = "User deleted."
    with get_db() as con:
        cams = con.execute("SELECT * FROM cameras WHERE active=1").fetchall()
        users = con.execute("SELECT id,username,role FROM users").fetchall()
        stats = con.execute(
            "SELECT cam_name, COUNT(*) as events, MAX(timestamp) as last "
            "FROM detections GROUP BY cam_name"
        ).fetchall()
        total = con.execute("SELECT COUNT(*) as c FROM detections").fetchone()["c"]
    return render_template("admin.html",
                           cams=cams, users=users, stats=stats,
                           total=total, msg=msg, user=session["user"])


# ── Video stream ──────────────────────────────────────────────────────────────
@app.route("/video/<int:cam_id>")
@login_required
def video(cam_id):
    def gen():
        while True:
            state = cameras.get(cam_id)
            if not state:
                time.sleep(0.1)
                continue
            frame = state.get("frame")
            if frame is None:
                time.sleep(0.03)
                continue
            _, buf = cv2.imencode(".jpg", frame, [cv2.IMWRITE_JPEG_QUALITY, 85])
            yield (b"--frame\r\nContent-Type: image/jpeg\r\n\r\n"
                   + buf.tobytes() + b"\r\n")
    return Response(gen(), mimetype="multipart/x-mixed-replace; boundary=frame")


# ── Recording ─────────────────────────────────────────────────────────────────
@app.route("/record/<int:cam_id>", methods=["POST"])
@login_required
def start_recording(cam_id):
    state = cameras.get(cam_id)
    if not state:
        return jsonify({"error": "Camera not found"}), 404
    if state.get("recording"):
        return jsonify({"error": "Already recording"}), 400

    with state["lock"]:
        frame = state["frame"]
    if frame is None:
        return jsonify({"error": "No frame available yet"}), 400

    fh, fw = frame.shape[:2]
    start_time = datetime.now()
    filename = os.path.join(RECORDINGS_DIR,
                            f"cam_{cam_id}_{start_time.strftime('%Y%m%d_%H%M%S')}.mp4")

    # Try avc1 first, fallback to mp4v
    writer = cv2.VideoWriter(filename, cv2.VideoWriter_fourcc(*'avc1'), 20.0, (fw, fh))
    if not writer.isOpened():
        writer = cv2.VideoWriter(filename, cv2.VideoWriter_fourcc(*'mp4v'), 20.0, (fw, fh))
    if not writer.isOpened():
        return jsonify({"error": "Failed to initialize video writer"}), 500

    state["recording"] = True
    state["writer"] = writer
    state["rec_start"] = start_time
    state["rec_filename"] = filename
    return jsonify({"status": "recording", "cam": state["name"]})


@app.route("/stop_record/<int:cam_id>", methods=["POST"])
@login_required
def stop_recording(cam_id):
    state = cameras.get(cam_id)
    if not state or not state.get("recording"):
        return jsonify({"error": "Not recording"}), 400

    state["recording"] = False
    writer = state.pop("writer", None)
    if writer:
        writer.release()
    state["writer"] = None

    filename = state.get("rec_filename", "")
    start_time = state.get("rec_start")
    end_time = datetime.now()

    with get_db() as con:
        con.execute(
            "INSERT INTO recordings (cam_id,cam_name,filename,start_time,end_time) VALUES (?,?,?,?,?)",
            (cam_id, state["name"], filename,
             start_time.strftime("%Y-%m-%d %H:%M:%S"),
             end_time.strftime("%Y-%m-%d %H:%M:%S"))
        )
    return jsonify({"status": "stopped", "file": os.path.basename(filename)})


# ── Clip extraction ───────────────────────────────────────────────────────────
@app.route("/clip/<int:cam_id>", methods=["POST"])
@login_required
def extract_clip(cam_id):
    data = request.json
    try:
        start_time = datetime.strptime(data["start"], "%Y-%m-%d %H:%M:%S")
        end_time   = datetime.strptime(data["end"],   "%Y-%m-%d %H:%M:%S")
    except Exception:
        return jsonify({"error": "Use format: YYYY-MM-DD HH:MM:SS"}), 400

    state = cameras.get(cam_id)
    if not state:
        return jsonify({"error": "Camera not found"}), 404

    frames = [(ts, f) for ts, f in list(state["buffer"]) if start_time <= ts <= end_time]
    if not frames:
        return jsonify({"error": "No frames in that time range"}), 404

    fh, fw = frames[0][1].shape[:2]
    filename = os.path.join(RECORDINGS_DIR,
                            f"clip_{cam_id}_{start_time.strftime('%H%M%S')}.mp4")

    writer = cv2.VideoWriter(filename, cv2.VideoWriter_fourcc(*'mp4v'), 20.0, (fw, fh))
    for _, f in frames:
        writer.write(f)
    writer.release()

    return jsonify({"status": "ok", "file": os.path.basename(filename),
                    "frames": len(frames), "download": f"/download/{filename}"})


@app.route("/download/<path:filename>")
@login_required
def download_file(filename):
    return send_file(filename, as_attachment=True)


# ── Export ────────────────────────────────────────────────────────────────────
@app.route("/export")
@login_required
def export():
    import io, openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    with get_db() as con:
        rows = con.execute("SELECT id,cam_name,details,timestamp FROM detections ORDER BY id ASC").fetchall()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Detection Log"
    thin = Side(style="thin", color="E5E7EB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.merge_cells("A1:D1")
    ws["A1"] = "PersonMonitor — Detection Report"
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="111827")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36
    ws.merge_cells("A2:D2")
    ws["A2"] = f"Generated: {datetime.now().strftime('%d %B %Y, %H:%M:%S')}"
    ws["A2"].font = Font(size=9, color="6B7280")
    ws["A2"].fill = PatternFill("solid", fgColor="F9FAFB")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.append([])
    headers = ["#", "Camera", "Detected Objects", "Timestamp"]
    ws.append(headers)
    for cell in ws[4]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F2937")
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    alt = PatternFill("solid", fgColor="F9FAFB")
    for i, r in enumerate(rows, 1):
        ws.append([i, r["cam_name"], r["details"], r["timestamp"]])
        for cell in ws[4 + i]:
            cell.border = border
            cell.fill = alt if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 22
    ws.freeze_panes = "A5"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(buf,
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename=detections_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"})


# ── AI ────────────────────────────────────────────────────────────────────────
@app.route("/ai", methods=["POST"])
@login_required
def ai_query():
    query = request.json.get("query", "").strip()
    if not query:
        return jsonify({"error": "Empty query"}), 400

    role = session.get("role")
    user = session.get("user")

    with get_db() as con:
        total = con.execute("SELECT COUNT(*) as c FROM detections").fetchone()["c"]
        recent = con.execute("SELECT cam_name,details,timestamp FROM detections ORDER BY id DESC LIMIT 50").fetchall()
        cam_stats = con.execute(
            "SELECT cam_name, COUNT(*) as events, MAX(timestamp) as last FROM detections GROUP BY cam_name"
        ).fetchall()
        if role == "admin":
            users = con.execute("SELECT username, role FROM users").fetchall()
            cams = con.execute("SELECT name, url FROM cameras WHERE active=1").fetchall()
            extra = f"Users: {[dict(u) for u in users]}\nCameras: {[dict(c) for c in cams]}\n"
        else:
            extra = ""

    recent_lines = "\n".join(f"  - {r['cam_name']}: {r['details']} at {r['timestamp']}" for r in recent)
    stat_lines   = "\n".join(f"  - {s['cam_name']}: {s['events']} events, last at {s['last']}" for s in cam_stats)

    restrictions = ("Full system access." if role == "admin" else
                    "Only detection data. No camera IPs, users, or admin settings.")

    prompt = f"""You are an AI assistant for PersonMonitor (real-time detection system).
User: {role} '{user}'. Access: {restrictions}

System data:
- Total events: {total}
{extra}
Camera stats:
{stat_lines}

Recent 50 detections:
{recent_lines}

Question: {query}

Answer concisely. If asking for a video clip, tell user to type:
/clip <cam_id> <YYYY-MM-DD HH:MM:SS> <YYYY-MM-DD HH:MM:SS>"""

    try:
        response = gemini.generate_content(prompt)
        return jsonify({"answer": response.text})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── SSE & Stats ───────────────────────────────────────────────────────────────
@app.route("/api/stats")
@login_required
def api_stats():
    with get_db() as con:
        recent = con.execute("SELECT cam_name,details,timestamp FROM detections ORDER BY id DESC LIMIT 30").fetchall()
        total = con.execute("SELECT COUNT(*) as c FROM detections").fetchone()["c"]
    live = {cid: {"detections": s["detections"], "time": s["time"],
                  "name": s["name"], "recording": s.get("recording", False)}
            for cid, s in cameras.items()}
    return jsonify({"total": total, "live": live, "recent": [dict(r) for r in recent]})


@app.route("/events")
@login_required
def events():
    def stream():
        last = {}
        while True:
            for cid, s in list(cameras.items()):
                key = f"{cid}:{','.join(s['detections'])}:{s['time']}"
                if last.get(cid) != key:
                    last[cid] = key
                    details = ", ".join(s["detections"]) if s["detections"] else "none"
                    rec = "true" if s.get("recording") else "false"
                    yield f"data: {cid}|{details}|{s['time']}|{s['name']}|{rec}\n\n"
            time.sleep(0.4)
    return Response(stream_with_context(stream()), mimetype="text/event-stream")


if __name__ == "__main__":
    init_db()
    load_cameras_from_db()
    port = int(os.getenv("PORT", 5000))
    app.run(debug=False, threaded=True, host="0.0.0.0", port=port)
